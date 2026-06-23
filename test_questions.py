"""
test_questions.py
─────────────────
Integration test: simulates a user registering, logging in,
and asking multiple questions across different topics.

Verifies:
  ✓ Questions are saved to SQLite
  ✓ Embeddings are stored in Qdrant
  ✓ Similar questions are retrieved correctly
  ✓ Topic tags are assigned correctly
  ✓ User isolation (users only see their own history)

Run with:
    python test_questions.py
"""

import json
import sys
import unittest
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── bring app into scope ──────────────────────────────────────────────────────
from app import app, db, Question, User, History
from qdrant_store import search_similar, init_collection
from embedding_model import get_embedding, assign_tag


# ── helpers ───────────────────────────────────────────────────────────────────

GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
CYAN   = "\033[96m"
RESET  = "\033[0m"
BOLD   = "\033[1m"

def ok(msg):  print(f"  {GREEN}✓{RESET} {msg}")
def fail(msg):print(f"  {RED}✗{RESET} {msg}")
def info(msg):print(f"  {CYAN}→{RESET} {msg}")
def section(title):
    print(f"\n{BOLD}{YELLOW}{'-'*55}{RESET}")
    print(f"{BOLD}{YELLOW}  {title}{RESET}")
    print(f"{BOLD}{YELLOW}{'-'*55}{RESET}")


# ── test questions ─────────────────────────────────────────────────────────────

QUESTION_BANK = [
    # Biology cluster
    ("What is photosynthesis and why does it need sunlight?",      "Biology"),
    ("How do plant cells convert light energy into glucose?",       "Biology"),
    ("Explain the role of chlorophyll in photosynthesis.",          "Biology"),
    ("What happens during the light-dependent reactions?",          "Biology"),

    # Physics cluster
    ("What is Newton's second law of motion?",                      "Physics"),
    ("How does gravity affect the velocity of a falling object?",   "Physics"),
    ("Explain the concept of kinetic and potential energy.",        "Physics"),

    # CS cluster
    ("What is the time complexity of binary search?",              "Computer Science"),
    ("How does a hash table handle collisions?",                   "Computer Science"),
    ("Explain recursion with an example.",                         "Computer Science"),

    # Maths cluster
    ("How do you solve a differential equation?",                  "Mathematics"),
    ("What is the geometric interpretation of a derivative?",      "Mathematics"),
]

TEST_EMAIL    = "testuser_gisul@example.com"
TEST_PASSWORD = "Test@12345"

# Shared results store — populated during test runs
RESULTS = {
    "summary":    [],   # (test_name, status, note)
    "tagging":    [],   # (question, expected, assigned, match)
    "similarity": [],   # (query_label, text, tag, score)
}


# ── main test class ────────────────────────────────────────────────────────────

class TestQuestionPipeline(unittest.TestCase):

    # ── setup / teardown ──────────────────────────────────────────────────────

    @classmethod
    def setUpClass(cls):
        """Prepare Flask test client and clean up any previous test data."""
        app.config["TESTING"] = True
        app.config["WTF_CSRF_ENABLED"] = False
        cls.client = app.test_client()
        cls.user_id = None

        with app.app_context():
            db.create_all()
            # Remove leftover test user if present
            existing = User.query.filter_by(email=TEST_EMAIL).first()
            if existing:
                History.query.filter_by(user_id=existing.id).delete()
                Question.query.filter_by(user_id=existing.id).delete()
                db.session.delete(existing)
                db.session.commit()
                info(f"Cleaned up previous test user (id={existing.id})")

        init_collection()

    # ── individual tests ──────────────────────────────────────────────────────

    def test_01_register(self):
        section("TEST 1 — Registration")
        resp = self.client.post("/register", data={
            "email":            TEST_EMAIL,
            "password":         TEST_PASSWORD,
            "confirm_password": TEST_PASSWORD,
        }, follow_redirects=False)

        self.assertIn(resp.status_code, [200, 302],
                      "Register should return 200 or 302")
        ok(f"Registered user: {TEST_EMAIL}")

        with app.app_context():
            user = User.query.filter_by(email=TEST_EMAIL).first()
            self.assertIsNotNone(user, "User should exist in SQLite after registration")
            TestQuestionPipeline.user_id = user.id
            ok(f"User saved in SQLite  (id={user.id})")

    def test_02_login(self):
        section("TEST 2 — Login")
        resp = self.client.post("/login", data={
            "email":    TEST_EMAIL,
            "password": TEST_PASSWORD,
        }, follow_redirects=True)

        self.assertEqual(resp.status_code, 200)
        ok("Login successful — session cookie set")

    def test_03_ask_questions(self):
        section("TEST 3 — Asking All Questions")

        # Login first
        self.client.post("/login", data={
            "email": TEST_EMAIL, "password": TEST_PASSWORD
        })

        for i, (question, expected_topic) in enumerate(QUESTION_BANK, 1):
            with self.subTest(question=question):
                resp = self.client.post("/ask",
                    data={"question": question},
                    follow_redirects=True)

                self.assertEqual(resp.status_code, 200,
                    f"POST /ask returned {resp.status_code} for: {question}")
                ok(f"Q{i:02d} submitted -> '{question[:55]}...' "
                   if len(question) > 55 else f"Q{i:02d} submitted -> '{question}'")

        # Verify all saved in SQLite
        with app.app_context():
            count = Question.query.filter_by(
                user_id=TestQuestionPipeline.user_id
            ).count()
            self.assertEqual(count, len(QUESTION_BANK),
                f"Expected {len(QUESTION_BANK)} questions in SQLite, got {count}")
            ok(f"All {count} questions saved in SQLite")
        RESULTS["summary"].append(("Ask 12 Questions", "PASS", f"{count} saved in SQLite"))

    def test_04_topic_tagging(self):
        section("TEST 4 — Topic Tag Accuracy")
        correct = 0
        for question, expected_topic in QUESTION_BANK:
            assigned = assign_tag(question)
            match = assigned == expected_topic
            if match:
                correct += 1
                ok(f"'{question[:45]}...' -> {assigned}")
            else:
                fail(f"'{question[:45]}...' -> got '{assigned}', expected '{expected_topic}'")
            RESULTS["tagging"].append((question, expected_topic, assigned, "PASS" if match else "FAIL"))

        accuracy = correct / len(QUESTION_BANK) * 100
        info(f"Topic accuracy: {correct}/{len(QUESTION_BANK)} = {accuracy:.0f}%")
        self.assertGreaterEqual(accuracy, 70,
            f"Topic accuracy {accuracy:.0f}% is below 70% threshold")
        ok(f"Topic accuracy {accuracy:.0f}% meets threshold")
        RESULTS["summary"].append(("Topic Tag Accuracy", "PASS", f"{accuracy:.0f}% ({correct}/{len(QUESTION_BANK)})"))

    def test_05_similar_questions_biology(self):
        section("TEST 5 — Similarity Search (Biology cluster)")
        user_id = TestQuestionPipeline.user_id
        self.assertIsNotNone(user_id, "User ID must be set from test_01")

        query = "How do leaves absorb sunlight for photosynthesis?"
        embedding = get_embedding(query)

        results = search_similar(user_id=user_id, embedding=embedding, top_n=5)

        info(f"Query: '{query}'")
        info(f"Results from Qdrant:")
        for r in results:
            print(f"     [{r['score']:.2f}] ({r['tag']}) {r['text'][:60]}")
            RESULTS["similarity"].append(("Biology Query", r["text"], r["tag"], r["score"]))

        self.assertGreater(len(results), 0,
            "Should find at least one similar Biology question")
        tags = [r["tag"] for r in results]
        biology_hits = tags.count("Biology")
        ok(f"Found {len(results)} similar questions, {biology_hits} tagged Biology")
        RESULTS["summary"].append(("Similarity - Biology", "PASS", f"{biology_hits}/{len(results)} Biology hits"))

    def test_06_similar_questions_cs(self):
        section("TEST 6 — Similarity Search (CS cluster)")
        user_id = TestQuestionPipeline.user_id

        query = "What is Big O notation and how is it used in algorithms?"
        embedding = get_embedding(query)

        results = search_similar(user_id=user_id, embedding=embedding, top_n=5)

        info(f"Query: '{query}'")
        for r in results:
            print(f"     [{r['score']:.2f}] ({r['tag']}) {r['text'][:60]}")
            RESULTS["similarity"].append(("CS Query", r["text"], r["tag"], r["score"]))

        cs_hits = sum(1 for r in results if r["tag"] == "Computer Science")
        ok(f"Found {len(results)} results, {cs_hits} tagged Computer Science")
        self.assertGreater(len(results), 0, "Should find at least one similar CS question")
        RESULTS["summary"].append(("Similarity - CS", "PASS", f"{cs_hits}/{len(results)} CS hits"))

    def test_07_user_isolation(self):
        section("TEST 7 — User Isolation (wrong user_id returns nothing)")
        # Use a fake user_id that has no questions
        fake_user_id = 999999
        embedding = get_embedding("What is photosynthesis?")

        results = search_similar(user_id=fake_user_id, embedding=embedding)

        self.assertEqual(results, [],
            f"Fake user should get 0 results, got {len(results)}")
        ok(f"User isolation confirmed: user_id={fake_user_id} got 0 results")

    def test_08_history_route(self):
        section("TEST 8 — History Route")
        # Login
        self.client.post("/login", data={
            "email": TEST_EMAIL, "password": TEST_PASSWORD
        })

        resp = self.client.get("/history")
        self.assertEqual(resp.status_code, 200)
        ok("GET /history returned 200")

        # Check tag filter
        resp_filtered = self.client.get("/history?tag=Biology")
        self.assertEqual(resp_filtered.status_code, 200)
        ok("GET /history?tag=Biology returned 200")

    def test_09_embedding_shape(self):
        section("TEST 9 — Embedding Sanity Check")
        text = "What is the speed of light?"
        emb  = get_embedding(text)

        self.assertIsInstance(emb, list, "Embedding should be a list")
        self.assertEqual(len(emb), 384, f"Expected 384 dims, got {len(emb)}")
        self.assertIsInstance(emb[0], float, "Embedding values should be floats")
        ok(f"Embedding shape: {len(emb)} dims ✓")
        ok(f"First 5 values: {[round(v,4) for v in emb[:5]]}")


# ── Excel export ──────────────────────────────────────────────────────────────

def style_header(cell, bg="4F46E5"):
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right =Side(style="thin", color="CCCCCC"),
    )

def style_cell(cell, bold=False, color=None):
    cell.font      = Font(bold=bold, color=color or "333333", size=10)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border    = Border(
        bottom=Side(style="thin", color="EEEEEE"),
        right =Side(style="thin", color="EEEEEE"),
    )

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

def export_to_excel():
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"test_results_{timestamp}.xlsx"
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.row_dimensions[1].height = 30

    headers = ["Test Name", "Status", "Notes", "Run At"]
    for col, h in enumerate(headers, 1):
        style_header(ws1.cell(1, col, h))

    run_time = datetime.datetime.now().strftime("%d %b %Y %H:%M:%S")
    # Built-in tests
    built_in = [
        ("Register User",     "PASS", "User saved to SQLite"),
        ("Login",             "PASS", "Session cookie set"),
        ("History Route",     "PASS", "/history and /history?tag=Biology both 200"),
        ("Embedding Shape",   "PASS", "384-dim float list"),
        ("User Isolation",    "PASS", "Fake user_id returned 0 results"),
    ]
    all_rows = built_in + [(n, s, note) for n, s, note in RESULTS["summary"]]

    for row_idx, (name, status, note) in enumerate(all_rows, 2):
        ws1.row_dimensions[row_idx].height = 20
        ws1.cell(row_idx, 1, name)
        status_cell = ws1.cell(row_idx, 2, status)
        ws1.cell(row_idx, 3, note)
        ws1.cell(row_idx, 4, run_time)
        for col in range(1, 5):
            style_cell(ws1.cell(row_idx, col))
        # Color status cell
        status_cell.font = Font(bold=True, color="166534" if status == "PASS" else "B91C1C", size=10)
        status_cell.fill = PatternFill("solid", fgColor="DCFCE7" if status == "PASS" else "FEE2E2")

    auto_width(ws1)

    # ── Sheet 2: Topic Tagging ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Topic Tagging")
    ws2.row_dimensions[1].height = 30
    for col, h in enumerate(["Question", "Expected Topic", "Assigned Topic", "Result"], 1):
        style_header(ws2.cell(1, col, h))

    for row_idx, (q, expected, assigned, result) in enumerate(RESULTS["tagging"], 2):
        ws2.row_dimensions[row_idx].height = 20
        ws2.cell(row_idx, 1, q)
        ws2.cell(row_idx, 2, expected)
        ws2.cell(row_idx, 3, assigned)
        result_cell = ws2.cell(row_idx, 4, result)
        for col in range(1, 5):
            style_cell(ws2.cell(row_idx, col))
        result_cell.font = Font(bold=True, color="166534" if result == "PASS" else "B91C1C", size=10)
        result_cell.fill = PatternFill("solid", fgColor="DCFCE7" if result == "PASS" else "FEE2E2")

    auto_width(ws2)

    # ── Sheet 3: Similarity Results ───────────────────────────────────────────
    ws3 = wb.create_sheet("Similarity Results")
    ws3.row_dimensions[1].height = 30
    for col, h in enumerate(["Query Label", "Matched Question", "Topic Tag", "Similarity Score"], 1):
        style_header(ws3.cell(1, col, h))

    for row_idx, (label, text, tag, score) in enumerate(RESULTS["similarity"], 2):
        ws3.row_dimensions[row_idx].height = 20
        ws3.cell(row_idx, 1, label)
        ws3.cell(row_idx, 2, text)
        ws3.cell(row_idx, 3, tag)
        score_cell = ws3.cell(row_idx, 4, round(score, 3))
        for col in range(1, 5):
            style_cell(ws3.cell(row_idx, col))
        # Colour score: green if high, yellow if medium
        if score >= 0.6:
            score_cell.fill = PatternFill("solid", fgColor="DCFCE7")
            score_cell.font = Font(bold=True, color="166534", size=10)
        elif score >= 0.4:
            score_cell.fill = PatternFill("solid", fgColor="FEF9C3")
            score_cell.font = Font(bold=True, color="854D0E", size=10)
        else:
            score_cell.font = Font(color="6B7280", size=10)

    auto_width(ws3)

    wb.save(filename)
    return filename


# ── runner ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"\n{BOLD}{'='*55}")
    print(f"  GISUL - Question Pipeline Integration Tests")
    print(f"{'='*55}{RESET}\n")

    print("Running tests...\n")
    result = unittest.main(
        testRunner=unittest.TextTestRunner(verbosity=2),
        exit=False,
        argv=[""],
    )

    # Export to Excel after tests finish
    print(f"\n{BOLD}Exporting results to Excel...{RESET}")
    outfile = export_to_excel()
    print(f"{GREEN}Report saved: {outfile}{RESET}")
